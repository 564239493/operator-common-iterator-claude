"""xlsx report parser for the project-local executer.

Mirrors the column-fuzzing strategy from
``operator-common-iterator/executer/report_parser.py`` — ATK report
column titles drift between versions, so we fuzzy-match against the
canonical Chinese / English names.  All parsing failures are captured
into ``TaskReportData.parse_error`` rather than raised, so result
extraction cannot abort the main execution flow.
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import (
    ComparisonRatio,
    ComparisonResult,
    ReportRecord,
    TaskReportData,
)

logger = logging.getLogger(__name__)


_RUN_RESULT_ALIASES = (
    "运行结果", "运行状态", "result", "run_result", "status", "Result",
    "测试结果", "结果",
)
_FAILURE_REASON_ALIASES = (
    "失败原因", "原因", "error", "fail_reason", "failure_reason",
    "失败信息", "error_message",
)
_CASE_JSON_ALIASES = (
    "用例JSON信息", "用例 JSON 信息", "用例json信息", "case_json",
    "用例JSON", "case_info", "用例信息", "case", "json",
)
_ID_ALIASES = (
    "用例ID", "用例 ID", "id", "ID", "case_id", "用例编号", "序号",
)

# fusion 精度对比比值列名别名（accuracy_load xlsx 实际列名落地需验，不覆盖则在此扩）。
_RATIO_ALIASES = {
    "max_re_ratio": (
        "max_re_ratio", "maxreratio", "最大相对误差", "max_re", "max相对误差比",
    ),
    "avg_re_ratio": (
        "avg_re_ratio", "avgreratio", "平均相对误差", "avg_re", "avg相对误差比",
    ),
    "root_mean_squared_ratio": (
        "root_mean_squared_ratio", "rmsratio", "rms", "均方根误差", "root_mean_squared",
    ),
}


def _norm(value: Any) -> str:
    """Lowercased, whitespace-collapsed string for fuzzy matching."""
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def _match_column(header: str, aliases: tuple[str, ...]) -> bool:
    target = _norm(header)
    for alias in aliases:
        if _norm(alias) == target:
            return True
    # Substring fallback for headers like "运行结果 (pass/fail)".
    return any(_norm(a) in target for a in aliases if len(_norm(a)) >= 2)


def _truthy_pass(value: Any) -> bool:
    """Interpret a ``run_result`` cell as a pass/fail boolean."""
    if value is None:
        return False
    text = _norm(value)
    if not text:
        return False
    if text in {
        "pass", "passed", "success", "ok", "成功", "通过", "1", "true", "yes", "y",
    }:
        return True
    if text in {
        "fail", "failed", "error", "fail_case", "失败", "未通过",
        "0", "false", "no", "n",
    }:
        return False
    # Anything containing 失败 / fail / error → fail; 通过 / pass → pass.
    if "失败" in text or "fail" in text or "error" in text:
        return False
    if "成功" in text or "通过" in text or "pass" in text:
        return True
    return False


def _find_latest_xlsx(report_dir: Path) -> Path | None:
    """Return the newest ``*.xlsx`` in ``report_dir`` by mtime, or ``None``."""
    if not report_dir.is_dir():
        return None
    candidates = sorted(
        report_dir.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def parse_xlsx_report(report_dir: Path) -> TaskReportData:
    """Parse the latest xlsx in ``report_dir`` into a :class:`TaskReportData`.

    Returns a ``TaskReportData`` with ``parse_error`` populated on failure
    instead of raising — execution flow must not abort on result
    extraction errors.
    """
    data = TaskReportData()

    if report_dir.is_file() and report_dir.suffix.lower() == ".xlsx":
        latest = report_dir
    else:
        latest = _find_latest_xlsx(report_dir)

    if latest is None:
        data.parse_error = "report 目录下未找到任何 .xlsx 文件"
        return data
    data.report_path = str(latest)

    try:
        wb = load_workbook(latest, read_only=True, data_only=True)
    except Exception as exc:
        data.parse_error = f"无法打开 xlsx: {exc}"
        return data

    try:
        ws = wb.active
        data.sheet_name = ws.title

        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            data.parse_error = "xlsx 为空, 无表头"
            return data

        col_for: dict[str, int] = {}
        for idx, header in enumerate(header_row):
            if _match_column(header or "", _ID_ALIASES):
                col_for.setdefault("id", idx)
            elif _match_column(header or "", _RUN_RESULT_ALIASES):
                col_for.setdefault("run_result", idx)
            elif _match_column(header or "", _FAILURE_REASON_ALIASES):
                col_for.setdefault("failure_reason", idx)
            elif _match_column(header or "", _CASE_JSON_ALIASES):
                col_for.setdefault("case_json", idx)

        for row in rows:
            if row is None:
                continue
            if all(
                cell is None or str(cell).strip() == "" for cell in row
            ):
                continue

            def _cell(key: str) -> Any:
                idx = col_for.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            case_json_raw = _cell("case_json")
            case_json_obj: dict[str, Any] | None = None
            if case_json_raw not in (None, ""):
                if isinstance(case_json_raw, dict):
                    case_json_obj = case_json_raw
                elif isinstance(case_json_raw, list):
                    case_json_obj = {"items": case_json_raw}
                else:
                    try:
                        parsed = json.loads(str(case_json_raw))
                        if isinstance(parsed, dict):
                            case_json_obj = parsed
                        elif isinstance(parsed, list):
                            case_json_obj = {"items": parsed}
                        else:
                            case_json_obj = {"value": parsed}
                    except (json.JSONDecodeError, TypeError, ValueError):
                        case_json_obj = {"raw": str(case_json_raw)}

            # Prefer the test case's DB id when present, fall back to
            # the xlsx id column.
            xlsx_id = (
                str(_cell("id")).strip()
                if _cell("id") is not None
                else ""
            ) or ""
            record_id = xlsx_id or None
            if isinstance(case_json_obj, dict):
                raw_cj_id = case_json_obj.get("id")
                if raw_cj_id is not None and str(raw_cj_id).strip():
                    record_id = str(raw_cj_id).strip()

            record = ReportRecord(
                id=record_id,
                run_result=(
                    str(_cell("run_result")).strip()
                    if _cell("run_result") is not None
                    else None
                ),
                failure_reason=(
                    str(_cell("failure_reason")).strip()
                    if _cell("failure_reason") is not None
                    else None
                ),
                case_json=case_json_obj,
                extra={
                    str(header_row[i] or f"col_{i}"): (
                        row[i] if i < len(row) else None
                    )
                    for i in range(len(header_row))
                    if i not in col_for.values()
                },
            )

            data.report_records.append(record)
            if _truthy_pass(record.run_result):
                data.passed += 1
            else:
                data.failed += 1
        data.record_count = len(data.report_records)
    except Exception as exc:
        logger.exception(
            "report_parser: xlsx parsing failed for %s", latest
        )
        data.parse_error = f"xlsx 解析失败: {exc}"
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return data


def _collect_ratio_values(
    records: list[ReportRecord], aliases: tuple[str, ...]
) -> list[float]:
    """Collect numeric values from any xlsx column fuzzy-matching ``aliases``.

    Scans every record's ``extra`` (catch-all for non-canonical columns) so
    per-case or summary ratio rows are both handled; non-numeric cells are
    skipped silently.  Mirrors the substring strategy of
    :func:`parse_xlsx_report`'s column matching.
    """
    values: list[float] = []
    for record in records:
        for key, val in (record.extra or {}).items():
            target = _norm(key)
            if not target:
                continue
            if not any(_norm(a) and _norm(a) in target for a in aliases):
                continue
            if val is None:
                continue
            try:
                values.append(float(str(val).strip()))
            except (ValueError, TypeError):
                continue
    return values


def parse_fusion_comparison(
    comparison_report_dir: Path,
    thresholds: dict[str, Any] | None = None,
) -> ComparisonResult:
    """Parse step4 ``accuracy_load`` xlsx into a :class:`ComparisonResult`.

    Record-only — does NOT set passed/failed or influence status.  Reuses
    :func:`parse_xlsx_report` for the xlsx open + row sweep, then pulls
    ``max_re_ratio / avg_re_ratio / root_mean_squared_ratio`` from the
    records' ``extra`` columns via fuzzy aliases.

    If the accuracy_load xlsx uses column names not covered by
    ``_RATIO_ALIASES``, the corresponding ``actual`` field stays ``None``
    — extend the alias tuples once the real column names are confirmed
    (see scheme.html 3.4.3 落地需验).
    """
    data = parse_xlsx_report(comparison_report_dir)
    actual = ComparisonRatio()
    records = data.report_records
    max_vals = _collect_ratio_values(records, _RATIO_ALIASES["max_re_ratio"])
    avg_vals = _collect_ratio_values(records, _RATIO_ALIASES["avg_re_ratio"])
    rms_vals = _collect_ratio_values(records, _RATIO_ALIASES["root_mean_squared_ratio"])
    if max_vals:
        actual.max_re_ratio = max(max_vals)
    if avg_vals:
        actual.avg_re_ratio = sum(avg_vals) / len(avg_vals)
    if rms_vals:
        actual.root_mean_squared_ratio = max(rms_vals)
    if data.parse_error:
        logger.warning(
            "parse_fusion_comparison: xlsx parse_error=%s; actual may be partial",
            data.parse_error,
        )
    return ComparisonResult(thresholds=thresholds, actual=actual)


__all__ = ["parse_fusion_comparison", "parse_xlsx_report"]
